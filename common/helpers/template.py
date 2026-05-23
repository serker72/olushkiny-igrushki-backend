import string
from datetime import datetime
from os import makedirs
from os.path import join
from typing import Any
from urllib.parse import ParseResult, parse_qs, urlencode, urlparse, urlunparse

import openpyxl
from fastapi import Request
from jinja2 import Template, pass_context
from jinja2.runtime import Context
from loguru import logger
from markupsafe import Markup
from starlette.datastructures import URL
from starlette.routing import NoMatchFound
from xlsxtpl.writerx import BookWriter

from common.exceptions import ConflictException, UnprocessableEntityException
from common.helpers.exception import get_traceback
from common.helpers.file import check_directory_exists
from common.models.base import Base
from common.schemas.models.settings import settings


def validate_notification_template(template_data: dict, entity: Base = None) -> list[None | datetime] | None:
    """Валидация данных шаблона уведомления"""
    if not template_data["sql_expression_list_recipients"]:
        raise UnprocessableEntityException(
            code="notification_template_sql_expression_list_recipients_empty",
            message_context={"code": template_data["code"]},
        )

    sql_expression = template_data["sql_expression_list_recipients"].lower()

    if template_data["is_check_entity_id"] is True and sql_expression.find(":entity_id") == -1:
        raise UnprocessableEntityException(
            code="notification_template_sql_expression_list_recipients_invalid",
            message_context={"code": template_data["code"]},
        )

    # Контроль наличия запрещенных операторов
    forbidden_statements = [
        "alter",
        "analize",
        "begin",
        "call",
        "checkpoint",
        "close",
        "cluster",
        "commit",
        "copy",
        "create",
        "deallocate",
        "declare",
        "delete",
        "discard",
        "do",
        "drop",
        "end",
        "execute",
        "explain",
        "fetch",
        "grant",
        "import",
        "insert",
        "listen",
        "load",
        "lock",
        "merge",
        "move",
        "notify",
        "prepare",
        "reassign",
        "refresh",
        "reindex",
        "release",
        "reset",
        "revoke",
        "rollback",
        "savepoint",
        "security",
        "set",
        "show",
        "start",
        "truncate",
        "unlisten",
        "update",
        "vacuum",
        "values",
    ]
    sql_expression_words = [word.strip(string.punctuation) for word in sql_expression.split()]
    forbidden_words = list(set(sql_expression_words) & set(forbidden_statements))

    if len(forbidden_words):
        raise UnprocessableEntityException(
            code="notification_template_sql_expression_list_recipients_forbidden_statements",
            message_context={"code": template_data["code"], "statements": ", ".join(forbidden_words)},
        )

    if template_data["is_delayed"] is True and (
        template_data["delayed_timeout_expressions"] is None
        or not isinstance(template_data["delayed_timeout_expressions"], list)
        or not len(template_data["delayed_timeout_expressions"])
    ):
        raise UnprocessableEntityException(
            code="notification_template_delayed_timeout_expressions_empty",
            message_context={"code": template_data["code"]},
        )

    if entity is None:
        return None

    started_on_times = []

    if template_data["is_current"] is True:
        started_on_times.append(None)

    if template_data["is_delayed"] is True:
        for timeout_expression in template_data["delayed_timeout_expressions"]:
            if timeout_expression.find("entity.") == -1:
                raise UnprocessableEntityException(
                    code="notification_template_delayed_timeout_expressions_invalid",
                    message_context={"code": template_data["code"], "expression": timeout_expression},
                )

            try:
                # timedelta используется в выражении
                if str(timeout_expression).find("timedelta") > -1:
                    from datetime import timedelta  # noqa: F401

                started_on = eval(timeout_expression)
            except Exception as e:
                trace = get_traceback(e)
                logger.error(trace)
                raise UnprocessableEntityException(
                    code="notification_template_delayed_timeout_expressions_not_eval",
                    message_context={"code": template_data["code"], "expression": timeout_expression, "error": str(e)},
                )

            started_on_times.append(started_on)

    return started_on_times


def render_and_save_template_xlsx(template_name: str, file_name: str, payloads: list[dict]):
    """Генерация и сохранение файла xlsx на основании шаблона"""
    writer = BookWriter(template_name)
    writer.jinja_env.globals.update(dir=dir, getattr=getattr)
    writer.render_book(payloads=payloads)
    writer.save(file_name)
    set_height_row(file_name)


def render_template_xlsx(template_name: str, file_path: str, file_name: str, payloads: list[dict]):
    """Генерация файла xlsx на основании шаблона"""
    if not check_directory_exists(file_path):
        try:
            makedirs(file_path, exist_ok=True)
        except OSError as e:
            raise ConflictException(code="directory_create_error", message_context={"name": file_path, "error": e})

    render_and_save_template_xlsx(template_name, join(file_path, file_name), payloads)


def set_height_row(file_path: str):
    """Установка высоты строки для отображения всего текста в ячейке"""
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    text_length = len(str(cell.value))
                    height = (text_length // 30 + 1) * 17
                    max_height = max(max_height, height)
            ws.row_dimensions[row[0].row].height = max_height
    wb.save(file_path)


def build_missing_route_url(request: Request, name: str, path_params: dict) -> str:
    """Построение полного URL для отсутствующих маршрутов"""
    parsed_base_url = urlparse(settings.backend_base_url)

    if name.startswith("http"):
        parsed_url = urlparse(name)
    else:
        _name = name.replace(".", "/")

        if _name.endswith("_page"):
            _name = _name[:-5]

        if _name.endswith("home"):
            _name = _name[:-4]

        # parsed_url.path = "/" + _name if not _name.startswith("/") else _name

        parsed_url = ParseResult(
            scheme=request.url.scheme,
            netloc=request.url.netloc,
            # path=request.url.path,
            path="/" + _name if not _name.startswith("/") else _name,
            params="",
            query=request.url.query,
            fragment=request.url.fragment,
        )

    path = (
        parsed_url.path + "/" + path_params.pop("path").lstrip("/")
        if path_params.get("path") is not None
        else parsed_url.path
    )
    params = str(path_params.pop("params", ""))

    is_ignored_query_params = path_params.pop("is_ignored_query_params", False)

    if len(path_params) > 0:
        query_params = parse_qs(parsed_url.query) if is_ignored_query_params is False else {}
        query_params.update(path_params)
        updated_query_string = urlencode(query_params, doseq=True)
    else:
        updated_query_string = ""

    updated_url = urlunparse(
        (
            parsed_base_url.scheme,
            parsed_url.netloc,
            path,
            params,
            updated_query_string,
            parsed_url.fragment,
        )
    )

    return str(updated_url)


@pass_context
def url_for(context: dict[str, Any], name: str, **path_params: Any) -> URL | str:
    """Построение полного URL для использования в шаблонах Jinja2"""
    request: Request = context["request"]
    filename = path_params.pop("filename", None)
    path = path_params.pop("path", None)

    if filename is not None or path is not None:
        path_params["path"] = filename or path

    try:
        # return request.url_for(name, **path_params)
        route_url = request.url_for(name, **path_params)
        parsed_route_url = urlparse(str(route_url))
        parsed_base_url = urlparse(settings.backend_base_url)

        updated_url = urlunparse(
            (
                parsed_base_url.scheme,
                parsed_route_url.netloc,
                parsed_route_url.path,
                parsed_route_url.params,
                parsed_route_url.query,
                parsed_route_url.fragment,
            )
        )

        return str(updated_url)
    except NoMatchFound:
        return build_missing_route_url(request, name, path_params)


@pass_context
def virtual_column(context: Context, value: str, item: Any):
    return Markup(Template(value).render({"item": item})).format()


@pass_context
def virtual_column_with_context(context: Context, value: str):
    return Markup(Template(value).render(context)).format()
